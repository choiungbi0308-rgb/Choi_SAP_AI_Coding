"""Generate ZMMR90810_SPEC.xlsx from spec data."""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# Color palette
CLR_HEADER_BG  = "1F4E79"
CLR_HEADER_FG  = "FFFFFF"
CLR_SUBHDR_BG  = "2E75B6"
CLR_SECTION_BG = "D6E4F0"
CLR_ROW_ALT    = "EBF3FB"
CLR_REQUIRED   = "C6EFCE"
CLR_WARN       = "FCE4D6"
CLR_GREEN_BG   = "E2EFDA"
CLR_RED_BG     = "FCE4D6"
CLR_BORDER     = "9DC3E6"

THIN = Side(style="thin", color=CLR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_font(bold=True, color=CLR_HEADER_FG, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def write_header_row(ws, row, values, bg=CLR_HEADER_BG, fg=CLR_HEADER_FG, height=20):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, color=fg, size=10)
        c.fill = fill(bg)
        c.alignment = wrap_align("center")
        c.border = BORDER

def write_data_row(ws, row, values, alt=False, row_colors=None, height=16):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        bg = CLR_ROW_ALT if alt else "FFFFFF"
        if row_colors and col in row_colors:
            bg = row_colors[col]
        c.fill = fill(bg)
        c.font = cell_font()
        c.alignment = wrap_align()
        c.border = BORDER

def write_section_title(ws, row, text, ncols, bg=CLR_SECTION_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, color="1F4E79", size=11)
    c.fill = fill(bg)
    c.alignment = wrap_align("left")
    c.border = BORDER
    ws.row_dimensions[row].height = 22


def sheet_overview(wb):
    ws = wb.active
    ws.title = "1.개요"
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [22, 45])

    overview = [
        ("프로그램 ID",    "ZMMR90810"),
        ("프로그램 제목",  "[GL] In & Out Daily Report"),
        ("트랜잭션 코드",  "ZMMR90810"),
        ("패키지",         "ZMM000"),
        ("메시지 클래스",  "ZMM01"),
        ("개발자",         "CHOIUB"),
        ("최초 작성일",    "2025-05-26"),
        ("Spec 작성일",    "2026-05-27"),
        ("SAP 모듈",       "MM — Inventory Management"),
    ]

    write_section_title(ws, 1, "프로그램 개요", 2)

    for i, (k, v) in enumerate(overview, 2):
        alt = (i % 2 == 0)
        c_k = ws.cell(row=i, column=1, value=k)
        c_v = ws.cell(row=i, column=2, value=v)
        for c in (c_k, c_v):
            c.fill = fill(CLR_ROW_ALT if alt else "FFFFFF")
            c.border = BORDER
            c.alignment = wrap_align()
        c_k.font = cell_font(bold=True)
        c_v.font = cell_font()
        ws.row_dimensions[i].height = 18


def sheet_selection(wb):
    ws = wb.create_sheet("2.선택화면")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [14, 14, 8, 22, 28, 30])

    headers = ["파라미터", "레이블", "필수", "기본값", "제약", "비즈니스 의미"]
    write_section_title(ws, 1, "선택 화면 (Selection Screen)", len(headers))
    write_header_row(ws, 2, headers)

    data = [
        ("S_WERKS", "플랜트",    "Yes", "—",           "단일값만 (NO INTERVALS)",      "조회 대상 플랜트"),
        ("S_CPUDT", "입력 날짜", "Yes", "D-1 ~ D-Day", "날짜 범위",                    "자재문서 CPU 일자 기준"),
        ("S_LGORT", "저장위치",  "No",  "—",           "단일값만",                      "플랜트 내 저장위치 한정 조회"),
        ("S_BWART", "이동유형",  "No",  "ZMMT90800 전체", "다단일값만",                   "INITIALIZATION 시 ZMMT90800 마스터에서 자동 채움"),
        ("S_MATNR", "자재번호",  "No",  "—",           "범위 가능",                    "특정 자재 한정 조회"),
    ]

    for i, row_data in enumerate(data, 3):
        alt = (i % 2 == 0)
        row_colors = {3: CLR_REQUIRED if row_data[2] == "Yes" else "FFFFFF"}
        if alt:
            row_colors = {col: CLR_ROW_ALT for col in range(1, 7)}
            if row_data[2] == "Yes":
                row_colors[3] = CLR_REQUIRED
        write_data_row(ws, i, row_data, alt=False, row_colors=row_colors, height=18)


def sheet_output(wb):
    ws = wb.create_sheet("3.출력필드")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [14, 18, 30, 20, 20])

    headers = ["필드명", "레이블", "원천", "비고", "색상/특이사항"]
    write_section_title(ws, 1, "출력 필드", len(headers))
    write_header_row(ws, 2, headers)

    data = [
        ("CPUDT",    "입력일자",    "MSEG.CPUDT_MKPF",                              "정렬 키 1",        "—"),
        ("WERKS",    "플랜트",      "MSEG.WERKS",                                   "정렬 키 2",        "—"),
        ("LGORT",    "저장위치",    "MSEG.LGORT",                                   "정렬 키 3",        "—"),
        ("MATNR",    "자재번호",    "MSEG.MATNR",                                   "정렬 키 4",        "더블클릭 → 드릴다운 팝업"),
        ("CURSTK",   "재고(D-Day)", "ZMMT80011 LABST+INSME+SPEME (ZDATE=CPUDT)",    "합계 출력",        "D-Day 종가 재고"),
        ("PRESTK",   "재고(D-1)",   "ZMMT80011 LABST+INSME+SPEME (ZDATE=CPUDT-1)", "합계 출력",        "전일 종가 재고"),
        ("INQTY",    "입고합계",    "SUM(MENGE) WHERE SHKZG='S'",                   "합계 출력",        "녹색 강조"),
        ("OUTQTY",   "출고합계",    "SUM(MENGE) WHERE SHKZG='H'",                   "합계 출력",        "빨간 강조"),
        ("MENGE",    "입출고합계",  "INQTY - OUTQTY",                               "합계 출력",        "순증감"),
        ("ZZCATEX1", "1차 카테고리","ZMMT90801.ZZCATEX",                            "이동유형 대분류",  "—"),
        ("ZZCATEX2", "2차 카테고리","ZMMT90802.ZZCATEX",                            "이동유형 중분류",  "—"),
        ("ZZCATEX3", "3차 카테고리","ZMMT90803.ZZCATEX",                            "이동유형 소분류",  "—"),
        ("BWART",    "이동유형",    "MSEG.BWART",                                   "—",               "—"),
        ("REMARK",   "비고",        "ZMMT90800.REMARK",                             "미등록: Etc.(확인필요)", "—"),
    ]

    color_map = {"INQTY": CLR_GREEN_BG, "OUTQTY": CLR_RED_BG}

    for i, row_data in enumerate(data, 3):
        alt = (i % 2 == 0)
        bg = color_map.get(row_data[0])
        if bg:
            row_colors = {col: bg for col in range(1, 6)}
        else:
            row_colors = {col: CLR_ROW_ALT for col in range(1, 6)} if alt else None
        write_data_row(ws, i, row_data, alt=False, row_colors=row_colors, height=18)


def sheet_bizrules(wb):
    ws = wb.create_sheet("4.비즈니스규칙")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [8, 45, 35])

    headers = ["#", "규칙", "영향"]
    write_section_title(ws, 1, "비즈니스 규칙 및 제약사항", len(headers))
    write_header_row(ws, 2, headers)

    data = [
        ("BR-01", "MARA.MATKL BETWEEN 'T001' AND 'T999' 범위 밖 자재는 제외",
                  "해당 자재는 경고 없이 누락됨"),
        ("BR-02", "ZMMT90800에 미등록 이동유형은 REMARK = 'Etc.(확인필요)'",
                  "정기적인 마스터 관리 필요"),
        ("BR-03", "D-Day/D-1 재고는 ZMMT80011 스냅샷 기반",
                  "당일 최신 재고와 차이 있을 수 있음"),
        ("BR-04", "S_WERKS는 단일값만 허용",
                  "복수 플랜트 동시 조회 불가"),
        ("BR-05", "ZMMT80011 데이터가 없는 날짜/자재는 CURSTK/PRESTK = 0",
                  "스냅샷 생성 배치 실패 시 공백 발생"),
    ]

    for i, row_data in enumerate(data, 3):
        alt = (i % 2 == 0)
        row_colors = {col: CLR_WARN for col in range(1, 4)}
        if not alt:
            row_colors = {col: "FFFFFF" for col in range(1, 4)}
        write_data_row(ws, i, row_data, alt=False, row_colors=row_colors, height=22)


def sheet_tables(wb):
    ws = wb.create_sheet("5.테이블목록")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [16, 10, 12, 30, 35])

    headers = ["테이블", "유형", "패키지", "주요 필드", "역할"]
    write_section_title(ws, 1, "테이블 및 구조체 목록", len(headers))
    write_header_row(ws, 2, headers)

    data = [
        ("MSEG",       "SAP 표준", "—",
         "MBLNR, MJAHR, CPUDT_MKPF, BWART, WERKS, LGORT, MATNR, MENGE, SHKZG",
         "자재문서 라인. ADBC Native SQL"),
        ("MARA",       "SAP 표준", "—",
         "MATKL",
         "자재 대분류 필터 (BETWEEN 'T001' AND 'T999')"),
        ("ZMMT90800",  "커스텀",   "ZMM000",
         "BWART, ZZCATE1, ZZCATE2, ZZCATE3, REMARK",
         "이동유형 카테고리 마스터"),
        ("ZMMT80011",  "커스텀",   "ZMM000",
         "ZDATE, WERKS, LGORT, MATNR, LABST, INSME, SPEME",
         "일별 재고 스냅샷"),
    ]

    sap_std_bg = "DEEAF1"
    custom_bg  = "E2EFDA"

    for i, row_data in enumerate(data, 3):
        bg = sap_std_bg if row_data[1] == "SAP 표준" else custom_bg
        row_colors = {col: bg for col in range(1, 6)}
        write_data_row(ws, i, row_data, alt=False, row_colors=row_colors, height=24)


def sheet_transport(wb):
    ws = wb.create_sheet("7.Transport대상")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [20, 12, 12, 35])

    headers = ["오브젝트", "유형", "패키지", "설명"]
    write_section_title(ws, 1, "Transport 대상 오브젝트", len(headers))
    write_header_row(ws, 2, headers)

    data = [
        ("ZMMR90810",      "PROG/P", "ZMM000", "메인 프로그램"),
        ("ZMMR90810_TOP",  "PROG/I", "ZMM000", "전역 선언 Include"),
        ("ZMMR90810_F01",  "PROG/I", "ZMM000", "FORM 루틴 Include"),
        ("ZMMR90810",      "TRAN/T", "ZMM000", "트랜잭션 코드"),
        ("ZMMT90800",      "TABD",   "ZMM000", "이동유형 카테고리 마스터"),
        ("ZMMT80011",      "TABD",   "ZMM000", "일별 재고 스냅샷 테이블"),
    ]

    prog_bg = "DEEAF1"
    tran_bg = "FFF2CC"
    tabd_bg = "E2EFDA"
    type_colors = {"PROG/P": prog_bg, "PROG/I": prog_bg,
                   "TRAN/T": tran_bg, "TABD": tabd_bg}

    for i, row_data in enumerate(data, 3):
        bg = type_colors.get(row_data[1], "FFFFFF")
        row_colors = {col: bg for col in range(1, 5)}
        write_data_row(ws, i, row_data, alt=False, row_colors=row_colors, height=18)


def main():
    wb = Workbook()
    sheet_overview(wb)
    sheet_selection(wb)
    sheet_output(wb)
    sheet_bizrules(wb)
    sheet_tables(wb)
    sheet_transport(wb)

    out_path = r"C:\AI-QA\scratch\tasks\ZMMR90810_SPEC.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
